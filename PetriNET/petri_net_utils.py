from __future__ import annotations

import datetime
import logging
import os
import random
from decimal import Decimal

from faker import Faker
from geopy import distance
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from .models import BusStop, City, Route

logger = logging.getLogger('PetriNetManager')

fake = Faker("ru_RU")

MAX_PASSENGERS_COUNT_FOR_RESPONSE = 10


def GetDataToCalculate(request_data_to_calculate: dict) -> dict:
    city_id = int(request_data_to_calculate['city_id'])
    DataToCalculate = {
        'city_id': city_id
    }

    # Получения маршрутов из бд
    routes = []
    for request_route in request_data_to_calculate['routes']:
        route = Route.objects.filter(city_id=city_id,
                                     id=request_route['id']).prefetch_related('busstop').first()
        # if not route:
        #     route = Route.objects.filter(city_id=city_id,
        #                                  name__icontains=request_route['name']).prefetch_related('busstop').first()
        if route:
            routes.append(route)
        else:
            logger.warning(f"Отсутствует маршрут, ID: {request_route['id']}")

    if not routes:
        raise Exception("Отсутствуют маршруты")
    DataToCalculate['routes'] = routes

    # Получение остановок и путей пассажиров
    busstops_directions = []
    busstops = BusStop.objects.filter(
        city_id=city_id,
        route__id__in=[route.id for route in routes]
    ).prefetch_related('route_set').distinct()

    for busstop in busstops:
        try:
            BSAddItem = {
                "busstop": busstop.id,
                "directions": {
                    # 3:5,  # 5 человек поедут на 3 ОП
                    # 4:3,  # 3 человека поедут на 4 ОП
                    # 7:2,  # 2 человека поедут на 7 ОП
                    # 0:12  # 12 человек поедут рандомно
                }
            }
            valid_bus_stops = BusStop.objects.filter(
                route__id__in=busstop.get_routes_ids()).values_list('id', flat=True).distinct()  # Пока делаем без пересадок
            busstop_from_request = request_data_to_calculate['busstops'].pop(str(busstop.id), None)
            if busstop_from_request:
                PassengersWithoutDirection = int(busstop_from_request.get('passengers_without_direction', 0))
                if PassengersWithoutDirection:
                    BSAddItem['directions'].update({
                        0: PassengersWithoutDirection
                    })
                for direction in busstop_from_request['directions']:
                    BusStopID = int(direction.get('busstop_id', 0))
                    PassengersCount = int(direction.get('passengers_count', 0))
                    if BusStopID and PassengersCount and busstops.filter(id=BusStopID).exists() and \
                       BusStopID != busstop.id and BusStopID in valid_bus_stops:
                        BSAddItem['directions'].update({
                            BusStopID: PassengersCount
                        })
                    else:
                        pass
            if BSAddItem['directions']:
                busstops_directions.append(BSAddItem)
        except Exception:
            logger.exception(f"Ошибка разбора данных остановки {busstop_from_request}")

    if not busstops_directions:
        raise Exception("Отсутствуют остановки")
    DataToCalculate['busstops'] = busstops
    DataToCalculate['busstops_directions'] = busstops_directions

    return DataToCalculate


def get_travel_time(latitude_start: float, longitude_start: float,
                  latitude_end: float, longitude_end: float, distance_km: int) -> int:
    # Написать функцию получения времени, за которое автобус проедет расстояние между остановками в секундах
    return int((distance_km / 40) * 3600)  # Пусть средняя скорость автобуса 40 км/ч


def get_travel_range(latitude_start: float, longitude_start: float,
                   latitude_end: float, longitude_end: float) -> int:
    """Нахождение расстояния между координатами в км"""
    return distance.distance((latitude_start, longitude_start), (latitude_end, longitude_end)).km


class PetriNet():
    passenger_time = 4  # Время входа или выхода пассажира (секунд)

    class Bus():
        """Автобус"""
        def __init__(self, route: Route, bus_id) -> None:
            # К какому маршруту относится
            self.route = route
            self.bus_id = bus_id
            # Вместимость
            self.capacity = self.route.tc.capacity
            # Координаты пути
            self.route_list = [self.serialize_route_point(point) for point in self.route.list_coord.copy()]
            self.bus_stop_ids = [bs.id for bs in self.route.busstop.all()]
            # Индекс остановки из пути
            self.bus_stop_index_now = 0
            # Пассажиры внутри
            self.passengers = []
            # Наверное сюда можно статистику добавить

        @property
        def get_rest_route(self) -> list[int]:
            """Получение списка id оставшихся ОП до конечной"""
            if self.ending_station():
                route_list = self.route_list.copy()
                route_list.reverse()
                return [point['bus_stop_id'] for point in route_list[0:]]
            return [point['bus_stop_id'] for point in self.route_list[self.bus_stop_index_now:]]

        def serialize_route_point(self, point: list):
            # Удобное хранение данных маршрута
            tolerance = Decimal('0.0001')

            bus_stop_id = 0
            lat, lon = point[0], point[1]
            
            # Сначала пробуем точное совпадение
            for bs in self.route.busstop.all():
                if lat == float(bs.latitude) and lon == float(bs.longitude):
                    bus_stop_id = bs.id
                    break
            
            # Если точное совпадение не найдено, ищем в радиусе tolerance
            if bus_stop_id == 0:
                nearby_stops = []
                for bs in self.route.busstop.all():
                    bs_lat = float(bs.latitude)
                    bs_lon = float(bs.longitude)
                    if (bs_lat >= lat - float(tolerance) and bs_lat <= lat + float(tolerance) and
                        bs_lon >= lon - float(tolerance) and bs_lon <= lon + float(tolerance)):
                        nearby_stops.append(bs)
                
                if nearby_stops:
                    if len(nearby_stops) > 1:
                        # Если найдено несколько остановок, выбираем ближайшую
                        closest_stop = min(nearby_stops, key=lambda stop: 
                                         get_travel_range(lat, lon, float(stop.latitude), float(stop.longitude)))
                        bus_stop_id = closest_stop.id
                    else:
                        bus_stop_id = nearby_stops[0].id
            
            if bus_stop_id == 0:
                raise Exception(f"Не удалось найти остановку для точки маршрута: {point}, {self.route.id}. Переформируйте маршрут.")

            return {
                "bus_stop_id": bus_stop_id,
                "latitude": lat,
                "longitude": lon,
                }

        def ending_station(self) -> bool:
            """Проверяет конечную станцию"""
            return True if self.bus_stop_index_now == len(self.route_list) - 1 else False
        
        def start_station(self) -> bool:
            """Проверяет начальную станцию"""
            return True if self.bus_stop_index_now == 0 else False

        def chech_of_travel_permit(self, busstops: 'PetriNet.BusStop', seconds_from_start: int) -> bool:
            """Проверяет что с текущей остановки достаточно давно выезжал автобус"""
            if self.route.id in busstops.routs_last_start_bus_time and \
                    (seconds_from_start - busstops.routs_last_start_bus_time[self.route.id]) < self.route.interval * 60:
                return False
            else:
                return True

        def get_travel_time(self) -> int:
            """Возвращает время, требуемое на дорогу до следующей остановки"""
            start_index = self.bus_stop_index_now
            end_index = start_index - 1 if self.ending_station() else start_index + 1
            distance_km = get_travel_range(
                self.route_list[start_index]["latitude"],
                self.route_list[start_index]["longitude"],
                self.route_list[end_index]["latitude"],
                self.route_list[end_index]["longitude"],
            )
            return get_travel_time(
                self.route_list[start_index]["latitude"],
                self.route_list[start_index]["longitude"],
                self.route_list[end_index]["latitude"],
                self.route_list[end_index]["longitude"],
                distance_km
                )

        def drive_to_next_bus_stop(self):
            """Передвигает автобус на следующиую остановку"""
            if self.ending_station():
                self.route_list.reverse()
                self.bus_stop_index_now = 1
            else:
                self.bus_stop_index_now += 1

        def get_action(self) -> dict:
            return {"Bus": [self]}

        def to_dict(self) -> dict:
            """Получение данных для отображения на сайте"""
            return {
                "bus_id": self.bus_id,
                "route_id": self.route.pk,
                "capacity": self.capacity,
                "bus_stop_id": self.route_list[self.bus_stop_index_now]["bus_stop_id"],
                "lat": self.route_list[self.bus_stop_index_now]["latitude"],
                "lng": self.route_list[self.bus_stop_index_now]["longitude"],
                "passengers": [pas.to_dict() for pas in self.passengers[0:MAX_PASSENGERS_COUNT_FOR_RESPONSE]],
                "passengers_count": len(self.passengers)
            }

    class BusStop():
        """Остановочный пункт"""
        def __init__(self, bus_stop: BusStop, passengers: list) -> None:
            self.bus_stop = bus_stop
            self.passengers = passengers
            # Время последнего отправления автобуса для каждого маршрута
            self.routs_last_start_bus_time = {}

        def set_last_start_bus_time(self, route_id: int, seconds_from_start: int):
            """Устанавливает остановке последнее время отправления для текущего маршрута"""
            self.routs_last_start_bus_time[route_id] = seconds_from_start

        def to_dict(self) -> dict:
            """Получение данных для отображения на сайте"""
            return {
                "id": self.bus_stop.pk,
                "lat": self.bus_stop.latitude,
                "lng": self.bus_stop.longitude,
                "passengers": [pas.to_dict() for pas in self.passengers[0:MAX_PASSENGERS_COUNT_FOR_RESPONSE]],
                "passengers_count": len(self.passengers)
            }

    class Passenger():
        """Пассажир"""
        def __init__(self, start_point: int, end_point: int) -> None:
            self.name = fake.first_name()
            self.start_bus_stop_id = start_point
            self.end_bus_stop_id = end_point

        def get_route_count(self, bus_stop_ids: list[int]) -> int:
            """Получение длительности пути пассажира (кол-ва остановок)"""
            start_index = bus_stop_ids.index(self.start_bus_stop_id)
            end_index = bus_stop_ids.index(self.end_bus_stop_id)
            if start_index >= end_index:
                raise Exception("Неправильное получение длительности пути пассажира")
            return end_index - start_index

        def to_dict(self) -> dict:
            """Получение данных для отображения на сайте"""
            return {
                "name": self.name,
                "start": self.start_bus_stop_id,
                "end": self.end_bus_stop_id
            }

    class TimeLine():
        """Порядок расчёта, класс работы с таймлайном"""
        def __init__(self, bus_stops_now: dict) -> None:
            self.timeline = {}
            # Указатель на автобусы из расчёта
            self.bus_stops_now = bus_stops_now
            # Данные для отправки на страницу расчёта
            self.data_to_response = []

        def add_timepoint(self, seconds_from_start: int, action: dict):
            """
            Функция добавление действиея в очередь

            seconds_from_start - Количество секунд со старта расчёта
            action - Данные в этот момент времени
            """
            if seconds_from_start in self.timeline:
                action_in_timeline = self.timeline[seconds_from_start]
                for key, value in action.items():
                    if key in action_in_timeline:
                        action_in_timeline[key].extend(value)
                    else:
                        action_in_timeline[key] = value
            else:
                self.timeline[seconds_from_start] = action

        def add_list_timepoints(self, list_timepoints: list) -> None:
            """
            Добавления списка действий в очередь
            """
            for action in list_timepoints:
                self.add_timepoint(seconds_from_start=action['seconds_from_start'],
                                   action=action['action'])

        def process_item_for_responce(self, item: dict) -> dict:
            item_for_responce = item.copy()
            for key, value in item.items():
                if isinstance(value, list):
                    item_for_responce[key] = [v.to_dict() for v in value]
                else:
                    item_for_responce[key] = value.to_dict()
            # Отправляем только остановки с пассажирами
            item_for_responce["BusStops"] = [bus_stop.to_dict() for bus_stop in self.bus_stops_now.values() if
                                             bus_stop.passengers]
            return item_for_responce

        def add_data_to_response(self, seconds_from_start: int, action: dict) -> None:
            """Добавляет действие для отрисовки"""
            self.data_to_response.append((seconds_from_start, self.process_item_for_responce(action)))

        def pop_first_timepoint(self) -> tuple[int, dict]:
            if not self.timeline:
                return None, None
            first_seconds_from_start, _ = self.get_first_timepoint()
            first_action = self.timeline.pop(first_seconds_from_start)
            self.data_to_response.append((first_seconds_from_start, self.process_item_for_responce(first_action)))
            return first_seconds_from_start, first_action

        def get_first_timepoint(self) -> tuple[int, dict]:
            list_actions = [key for key in self.timeline.keys()]
            list_actions.sort()
            first_key = list_actions[0]
            return first_key, self.timeline[first_key]

    def __init__(self, data_to_calculate: dict = {}) -> None:
        self.data_to_calculate = data_to_calculate
        self.routes = data_to_calculate['routes']
        self.data_to_report = {'routes': {route.id: {'route': route,
                                                     'average_passengers_stops_count': [0, 0],
                                                     'average_fullness': [0, 0],
                                                     'completed_trips': 0,  # Количество завершённых рейсов (достижений конечной)
                                                     } for route in self.routes}}
        # Список объектов остановок с пассажирами
        self.busstops: dict[int, PetriNet.BusStop] = {}
        self.busstops_cached: dict[int, BusStop] = {busstop.id: busstop for busstop in data_to_calculate['busstops']}
        self.timeline = self.TimeLine(self.busstops)
        self.init_action()
        # Наверное переделать Имитацию работы онлайн с 400мс. на относительную скорость движения между actions

    def init_action(self):
        """Создание объектов для расчёта (пассажиров и автобусов)"""
        add_timepoints = []

        for route in self.routes:
            time = 0
            if not route.amount or not route.tc:
                logger.warning(f"На маршруте {route.id} не указаны автобусы, маршрут не будет учитываться в расчёте")
            else:
                for bus_id in range(route.amount):
                    add_timepoints.append({
                        "seconds_from_start": time,
                        "action": {
                            "Bus": [
                                self.Bus(route, bus_id + 1),
                            ]
                        }
                    })
                    time += route.interval * 60

        if not add_timepoints:
            raise Exception("Отсутствуют автобусы на маршрутах")

        for busstops_direction in self.data_to_calculate['busstops_directions']:
            bus_stop = self.busstops_cached[busstops_direction['busstop']]
            valid_bus_stops = list(BusStop.objects.filter(
                route__id__in=bus_stop.get_routes_ids(),
                id__in=list(self.busstops_cached.keys())).values_list('id', flat=True).distinct())
            if busstops_direction['busstop'] in valid_bus_stops:
                valid_bus_stops.remove(busstops_direction['busstop'])
            passengers = []
            for direction, count in busstops_direction['directions'].items():
                if direction == 0:
                    passengers.extend(
                        self.Passenger(bus_stop.id,
                                       random.choice(valid_bus_stops)) for pas in range(count)
                                       )
                else:
                    passengers.extend(self.Passenger(bus_stop.id, direction) for pas in range(count))
            self.busstops.update({bus_stop.id: self.BusStop(bus_stop, passengers)})

        if not self.busstops:
            raise Exception("Отсутствуют пассажиры")

        for busstop in self.data_to_calculate['busstops']:
            if busstop.id not in self.busstops:
                self.busstops.update({busstop.id: self.BusStop(busstop, [])})

        self.timeline.add_list_timepoints(add_timepoints)

    def Calculation(self):
        """Модуль расчёта"""
        # Получаем первый таймпоинт
        this_seconds_from_start: int | None
        this_action: dict | None
        this_seconds_from_start, this_action = self.timeline.pop_first_timepoint()
        # Проходим по всем таймпоинтам
        while this_seconds_from_start or this_action:
            # Проверяем все автобусы в таймпоинте
            bus: PetriNet.Bus
            for bus in this_action["Bus"].copy():
                # Сколько временя заняло действие
                time_delta: int = 0
                # id текущей остановки автобуса
                bus_stop_id_now: int = bus.route_list[bus.bus_stop_index_now]["bus_stop_id"]
                # Сначала высаживаются из автобуса
                pas: PetriNet.Passenger
                for pas in bus.passengers.copy():
                    # Если конечная точка пассажира совпадает с текущей автобуса
                    if pas.end_bus_stop_id == bus_stop_id_now:
                        # Пассажир прибыл в место назначения
                        bus.passengers.remove(pas)
                        time_delta += self.passenger_time
                # Добавить таймпоинт после высадки людей
                if time_delta:
                    this_seconds_from_start += time_delta
                    time_delta = 0
                    self.timeline.add_data_to_response(this_seconds_from_start, bus.get_action())
                # Заходят пассажиры, которые могут доехать до своей остановки
                for pas in self.busstops[bus_stop_id_now].passengers.copy():
                    # Если конечная точка пассажира есть в оставшемся пути автобуса (до конечной),
                    # и места в автобусе ещё есть
                    if pas.end_bus_stop_id in bus.get_rest_route and len(bus.passengers) < bus.capacity:
                        # Пассажир уходит с остановки
                        self.busstops[bus_stop_id_now].passengers.remove(pas)
                        # Садится в автобус
                        bus.passengers.append(pas)
                        # Считается средняя длительность пути пассажиров
                        self.data_to_report['routes'][bus.route.id]['average_passengers_stops_count'][0] +=\
                            pas.get_route_count(bus.get_rest_route)
                        self.data_to_report['routes'][bus.route.id]['average_passengers_stops_count'][1] += 1
                        # Это занимает некоторое время
                        time_delta += self.passenger_time
                # Добавить таймпоинт после посадки людей
                if time_delta:
                    this_seconds_from_start += time_delta
                    time_delta = 0
                    self.timeline.add_data_to_response(this_seconds_from_start, bus.get_action())

                def exists_pass_in_bus_path(bus_stop_ids: list[int]) -> bool:
                    """Проверяем есть ли пассажиры на пути автобуса"""
                    stop_intersections: set[int] = set(bus_stop_ids) & \
                        {busstop.bus_stop.id for busstop in self.busstops.values() if busstop.passengers}
                    return bool(stop_intersections)

                # Автобус отправляется на следующую остановку,
                # если она конечная: если есть пассажиры на его пути или в нём едем дальше, иначе останавливаемся
                if (bus.ending_station() or bus.start_station()) and not bus.passengers and not exists_pass_in_bus_path(bus.bus_stop_ids):
                    # В этот момент можно у маркера отключить анимацию
                    # Автобус завершил работу - на маршруте больше нет пассажиров
                    pass
                else:
                    # При достижении начальной остановки, если за последние route.interval минут выезжал автобус
                    if bus.ending_station() and not bus.chech_of_travel_permit(self.busstops[bus_stop_id_now],
                                                                               this_seconds_from_start):
                        # Этот автобус ждёт route.interval минут
                        time_delta += bus.route.interval * 60
                    else:
                        # Устанавливаем остановке последнее время отправления для текущего маршрута
                        self.busstops[bus_stop_id_now].set_last_start_bus_time(bus.route.id, this_seconds_from_start)
                        # Добавляем время пути до следующей остановки
                        time_delta += bus.get_travel_time()
                        # Считаем среднюю наполненность
                        self.data_to_report['routes'][bus.route.id]['average_fullness'][0] += len(bus.passengers)
                        self.data_to_report['routes'][bus.route.id]['average_fullness'][1] += 1
                        # Считаем завершённый рейс когда автобус достигает конечной остановки
                        if bus.ending_station():
                            self.data_to_report['routes'][bus.route.id]['completed_trips'] += 1
                        # Передвигаем автобус
                        bus.drive_to_next_bus_stop()
                    this_action["Bus"].remove(bus)
                    # Добавляем следующий таймпоинт в таймлайн
                    self.timeline.add_timepoint(this_seconds_from_start + time_delta, bus.get_action())
            this_seconds_from_start, this_action = self.timeline.pop_first_timepoint()
        self.timeline.data_to_response.sort(key=lambda i: i[0])
        return self.timeline.data_to_response

    def CreateDataToReport(self) -> dict:
        """Собирает данные для отчёта"""
        data_to_report = {}
        data_to_report['city_name'] = City.objects.get(id=self.data_to_calculate['city_id']).name
        data_to_report['data'] = str(datetime.datetime.now().isoformat(sep='_', timespec='seconds')).replace(':', '-')
        data_to_report['bus_stops'] = []
        for bus_stop in self.data_to_calculate['busstops']:
            bus_add = {}
            for bus_in_calculate in self.timeline.data_to_response[0][1]['BusStops']:
                if bus_in_calculate['id'] == bus_stop.id:
                    bus_add['bus_name'] = bus_stop.name
                    bus_add['passengers_count'] = bus_in_calculate['passengers_count']
                    break
            if bus_add:
                for timepoint in self.timeline.data_to_response:
                    for bus_in_calculate in timepoint[1]['BusStops']:
                        if bus_in_calculate['id'] == bus_stop.id:
                            bus_add['max_waiting_time'] = timepoint[0]
                            break
                    else:
                        break
                bus_add['max_waiting_time'] = int(bus_add['max_waiting_time'] / 60)
                bus_add['routes_count'] = len(bus_stop.route_set.all())
                data_to_report['bus_stops'].append(bus_add)
        results_add = {
            'bus_name': 'Итоги',
            'passengers_count': 0,
            'max_waiting_time': [],
            'routes_count': ''
        }
        for bus_stop in data_to_report['bus_stops']:
            results_add['passengers_count'] += bus_stop['passengers_count']
            results_add['max_waiting_time'].append(bus_stop['max_waiting_time'])
        # Среднее время ожидания
        results_add['max_waiting_time'] = int(sum(results_add['max_waiting_time']) /
                                              len(results_add['max_waiting_time']))
        data_to_report['bus_stops'].append(results_add)
        # Формировать цвет время ожидания автобуса относительно среднего
        # bus_add['color'] = 'white'
        data_to_report['routes'] = []
        for route in self.data_to_report['routes'].values():
            add_route = route.copy()
            add_route.pop('route')
            add_route['name'] = route['route'].name
            TC = route['route'].tc
            add_route['TC'] = f'{TC.name}, {TC.capacity}' if TC else ''
            add_route['interval'] = route['route'].interval
            add_route['average_passengers_stops_count'] = round(route['average_passengers_stops_count'][0] /
                                                                (route['average_passengers_stops_count'][1] or 1), 2)
            add_route['average_fullness'] = str(round((route['average_fullness'][0] /
                                                       (route['average_fullness'][1] or 1) /
                                                       (TC.capacity or 1)) * 100, 2)) + '%'
            add_route['bus_stop_count'] = len(route['route'].busstop.all())
            # Расчёт протяжённости маршрута по координатам в порядке следования
            add_route['route_length'] = 0
            list_coord = route['route'].list_coord
            if list_coord and len(list_coord) > 1:
                for i in range(len(list_coord) - 1):
                    add_route['route_length'] += get_travel_range(
                        list_coord[i][0],
                        list_coord[i][1],
                        list_coord[i + 1][0],
                        list_coord[i + 1][1],
                    )
            add_route['route_length'] = round(add_route['route_length'], 2)
            add_route['TC_count'] = route['route'].amount
            add_route['trips_count'] = route['completed_trips']
            data_to_report['routes'].append(add_route)
        # Добавляем суммарное количество поездок (завершённых рейсов)
        data_to_report['total_trips_count'] = sum(route['trips_count'] for route in data_to_report['routes'])
        return data_to_report


def CreateResponseFile(data_to_report: dict) -> str:
    city_name = data_to_report.get('city_name', '')
    date = data_to_report.get('data', '')

    path = 'media/reports/'
    file_name = f'report_{city_name}_{date}.xlsx'
    file_path = path + file_name
    if os.path.isfile(file_path):
        return file_path

    # Создаем новый документ Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    # Стили
    header_font = Font(name='Times New Roman', size=14, bold=True)
    normal_font = Font(name='Times New Roman', size=14)
    title_font = Font(name='Times New Roman', size=16, bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    bus_stops = data_to_report.get('bus_stops', [])
    routes = data_to_report.get('routes', [])

    current_row = 1

    # Заголовок отчёта
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
    title_cell = ws.cell(row=current_row, column=1,
                         value=f'Результат расчёта нагрузки на транспортную сеть с использованием маршрутов: {", ".join([route["name"] for route in routes])}')
    title_cell.font = title_font
    title_cell.alignment = left_alignment
    ws.row_dimensions[current_row].height = 40
    current_row += 2

    # Информация о населённом пункте и дате
    ws.cell(row=current_row, column=1, value=f"Населённый пункт: {city_name}").font = normal_font
    current_row += 1
    ws.cell(row=current_row, column=1, value=f"Дата: {date}").font = normal_font
    current_row += 2

    # Таблица автобусных остановок
    if bus_stops:
        ws.cell(row=current_row, column=1, value='Автобусные остановки:').font = header_font
        current_row += 1

        bus_stop_headers = ['Остановка', 'Количество пассажиров',
                           'Максимальное время ожидания автобуса, мин.', 'Количество маршрутов, шт.']

        for col, header in enumerate(bus_stop_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1

        for stop in bus_stops:
            row_data = [
                stop.get('bus_name', ''),
                stop.get('passengers_count', ''),
                stop.get('max_waiting_time', ''),
                stop.get('routes_count', '')
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = normal_font
                cell.alignment = center_alignment
                cell.border = thin_border
            current_row += 1

        current_row += 1

    # Таблица маршрутов
    if routes:
        ws.cell(row=current_row, column=1, value='Маршруты:').font = header_font
        current_row += 1

        route_headers = [
            'Маршрут',
            'Тип ТС, Название, вместимость',
            'Интервал движения, мин.',
            'Средняя длительность пути, кол-во ОП',
            'Средняя наполненность, %',
            'Количество остановок',
            'Протяжённость, км.',
            'Кол-во автобусов',
            'Кол-во поездок'
        ]

        for col, header in enumerate(route_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        current_row += 1

        for route in routes:
            row_data = [
                route.get('name', ''),
                route.get('TC', ''),
                route.get('interval', ''),
                route.get('average_passengers_stops_count', ''),
                route.get('average_fullness', ''),
                route.get('bus_stop_count', ''),
                route.get('route_length', ''),
                route.get('TC_count', ''),
                route.get('trips_count', '')
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = normal_font
                cell.alignment = center_alignment
                cell.border = thin_border
            current_row += 1

        # Итоговая строка
        total_trips = data_to_report.get('total_trips_count', 0)
        for col in range(1, 10):
            cell = ws.cell(row=current_row, column=col, value='')
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        ws.cell(row=current_row, column=1, value='Итого').font = header_font
        ws.cell(row=current_row, column=9, value=total_trips).font = header_font

    # Автоматическая ширина столбцов на основе содержимого
    def get_column_width(text, font_size=14):
        """Расчёт ширины столбца для Times New Roman 14"""
        if text is None:
            return 0
        text = str(text)
        # Коэффициент для Times New Roman 14 (примерно 1.2 символа на единицу ширины)
        char_width = font_size * 0.15
        # Учитываем переносы строк
        lines = text.split('\n')
        max_line_length = max(len(line) for line in lines) if lines else 0
        return max_line_length * char_width + 2  # +2 для отступов

    for col_idx in range(1, ws.max_column + 1):
        max_width = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell_width = get_column_width(cell.value)
                max_width = max(max_width, cell_width)
        # Ограничиваем максимальную ширину для читаемости
        max_width = min(max_width, 50)
        max_width = max(max_width, 10)  # Минимальная ширина
        ws.column_dimensions[column_letter].width = max_width

    # Высота строк заголовков таблиц
    for row_idx in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        if cell.font and cell.font.bold and cell.border and cell.border.top.style:
            ws.row_dimensions[row_idx].height = 45

    if not os.path.exists(path):
        os.makedirs(path)
    # Сохраняем файл локально на сервере
    wb.save(file_path)

    return file_path

# Проверка что на маршруте 2 и более остановок
